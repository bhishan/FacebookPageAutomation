import facebook
import openpyxl
import glob
import os
from bs4 import BeautifulSoup
from urllib.request import urlopen

cfg = {
    "page_id"      : "your page id",
    "access_token" : "your page access token"   
    }

def scrape_content():
  wb = openpyxl.Workbook()
  ws = wb.active    
  soup = BeautifulSoup(urlopen('http://planetpython.org').read())
  article_content_divs = soup.find_all('h4')
  for each_content_div in article_content_divs:
    content = each_content_div.find('a')
    content_url = content['href']
    content_title = content.text
    ws.append([content_url, content_title])
    wb.save('fb_content.xlsx')
  wb.close()

def get_content():
  wb = openpyxl.load_workbook("fb_content.xlsx", read_only=False)
  ws = wb.active
  try:
    with open('sheet_row_pointer.txt', 'r') as row_pointer:
      last_content = int(row_pointer.read())
  except FileNotFoundError:
    last_content = 2
  if last_content == ws.max_row:
    wb.close()
    try:
      os.remove('sheet_row_pointer.txt')
    except:
      pass
    scrape_content()
    get_content()
  req_contents = ws.cell(row=last_content,column=1).value, ws.cell(row=last_content,column=2).value
  with open('sheet_row_pointer.txt', 'w') as row_pointer:
    row_pointer.write(str(last_content + 1))
  return req_contents

def get_api(cfg):
  graph = facebook.GraphAPI(cfg['access_token'])
  return graph

def main():
  api = get_api(cfg)
  if not glob.glob('*.xlsx'):
    scrape_content()
  content_url, content_title = get_content()
  msg = content_title + '\n' + content_url
  status = api.put_wall_post(msg)

if __name__ == "__main__":
  main()

